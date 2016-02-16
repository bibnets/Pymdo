#!/usr/bin/python
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename = 'test.xlsx')
ws1 = wb.worksheets[0]

c1 = ws1.cell(row=1,column=1).value
print c1
