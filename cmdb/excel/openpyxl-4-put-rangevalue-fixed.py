#!/usr/bin/python
# openpyxl operate excel rows

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

wb = load_workbook(filename = 'test.xlsx')
ws1 = wb.worksheets[0]
ws2 = ExcelWriter(workbook = wb)

for row in ws1.rows:
    for col in row:
        ns_name = col.value
        ws1.cell(column=2, row=2).value = col.value

ws2.save(filename = 'test2.xlsx')