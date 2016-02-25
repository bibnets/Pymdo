#!/usr/bin/python
# openpyxl operate excel rows

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

wb = load_workbook(filename = 'test.xlsx')
ws1 = wb.worksheets[0]
ws2 = ExcelWriter(workbook = wb)

for index,row in enumerate(ws1.rows):
    row_num=index+1
    for col in row:
        ws1.cell(column=2, row=row_num).value = col.value

ws2.save(filename = 'test2.xlsx')