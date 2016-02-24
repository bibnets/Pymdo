#!/usr/bin/python
# openpyxl operate excel rows

from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename = 'test.xlsx')
ws1 = wb.worksheets[0]


for row_num in range(1,30):
    ws1_value= ws1.cell(row=row_num,column=1).value
    if ws1_value != None:
        print ws1_value
    else:
        break

